#!/usr/bin/env python3.13
"""
ZimLivestock — Financial Model  (v3.0, June 2026 — scaling B2B-SaaS platform)

A 5-YEAR projection for a SCALING B2B-SaaS PLATFORM. Auction houses onboard as
isolated (RLS) tenants via a self-serve onboarding wizard and pay a one-off
onboarding fee + a monthly platform subscription + a thin 0.75% take on settled
GMV. The platform diversifies into B2B2C consumer revenue via TRANSPORT
(delivery) over time. B2B today, B2B2C tomorrow.

This supersedes the earlier v2.x model. The old bootstrap framing is dropped:
this is a small team (~9-13 people by Year 5), largely SELF-FUNDED with a prudent
2-3 month working-capital buffer (NOT external equity), and the founder draws a
salary that sits inside payroll. Growth rides on HOUSE COUNT + TRANSPORT, not on
pushing adoption (adoption is held field-honest, not used as a growth lever).

Scale: MODERATE — ~20 houses (about 1/3 of the ~40-60 house market) over 5
years, ramping 5 -> 8 -> 12 -> 16 -> 20 live at year-end, anchors-first.

Revenue lines (four):
- Onboarding fee (one-off):  Tier A US$3,500 · Tier B US$2,500 · Tier C US$1,500
                             · Pilot US$1,000 (credited toward the tier fee on conversion).
- Platform subscription (monthly): Tier A US$1,500 · Tier B US$1,200 · Tier C US$900
                             · Pilot US$1,000/mo (90 days, then converts to signed tier).
- Transaction take:          0.75% of settled GMV, on top of Paynow's fee.
- Transport (consumer line): buyer quote US$15 base + US$0.35/km, capped US$250.
                             Platform keeps ONLY the flat US$15 booking leg
                             (US$12.00 net after ~US$3 processing). Attach-rate
                             ramps 5% -> 9% -> 14% -> 19% -> 24% (Y1..Y5).

Cost lines (four): infrastructure, payroll, pass-through, other (BD/travel/mktg).

Outputs an .xlsx with 6 tabs:
  0. README          — what's in here + honest headline outputs
  1. Assumptions     — every input, editable
  2. House ramp      — houses live by year, tier mix, GMV onto rails
  3. Revenue         — annual revenue per line (4 lines), totaled
  4. Costs           — annual costs per line (4 lines), totaled
  5. P&L summary     — year-by-year income statement + drivers + honest caveats

Layout note: the canonical model is ANNUAL over 5 years, so tabs 3/4/5 use a
year-per-column layout (Y1..Y5 + 5yr-total) rather than the old month-per-column
grid. All figures are emitted directly from the v3.0 canon (no recomputation).

Run:  python3.13 financial-model-build.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# THE MODEL — single source of truth (v3.0 canon; mirrors gtm-strategy.md)
# ----------------------------------------------------------------------------
YEARS = 5
TAKE_RATE = 0.0075   # 0.75% of settled GMV, on top of Paynow's fee

# Houses live at year-end (anchors-first ramp).
HOUSES_LIVE = [5, 8, 12, 16, 20]

# Onboarding mix by year (net adds): (Tier A, Tier B, Tier C).
ONBOARDING_ADDS = [(3, 1, 1), (1, 1, 1), (1, 2, 1), (1, 2, 1), (1, 2, 1)]
# Cumulative tier mix (EOY): Y1 3A/1B/1C -> Y5 7A/8B/5C.
TIER_MIX_CUM = [(3, 1, 1), (4, 2, 2), (5, 4, 3), (6, 6, 4), (7, 8, 5)]

# Pricing (canonical).
ONBOARD_FEE = {"A": 3500, "B": 2500, "C": 1500, "Pilot": 1000}
SUBSCRIPTION = {"A": 1500, "B": 1200, "C": 900, "Pilot": 1000}

# Transport (consumer line).
TRANSPORT_BASE = 15.0          # buyer-facing flat booking leg
TRANSPORT_PER_KM = 0.35        # transporter keeps the distance haul
TRANSPORT_CAP = 250.0          # quote cap
TRANSPORT_NET = 12.00          # platform net after ~US$3 processing
ATTACH_RATE = [0.05, 0.09, 0.14, 0.19, 0.24]   # Y1..Y5

# Drivers (canonical).
BLENDED_ADOPTION = [0.105, 0.115, 0.123, 0.130, 0.137]   # field-honest, below ~15% ceiling
AVG_DIGITAL_TICKET = 650
DIGITAL_TRANSACTIONS = [1376, 3790, 5667, 8170, 10999]

# Revenue by line (US$, exact from canon).
REV_ONBOARDING   = [14500, 7500, 10000, 10000, 10000]      # 5yr 52,000
REV_SUBSCRIPTION = [39600, 100800, 151200, 208800, 266400] # 5yr 766,800
REV_TAKE         = [6710, 18474, 27625, 39831, 53618]       # 5yr 146,258
REV_TRANSPORT    = [826, 4093, 9521, 18628, 31677]          # 5yr 64,745
REV_TOTAL        = [61636, 130867, 198346, 277259, 361695]  # 5yr 1,029,803

# Costs by line (US$, exact from canon).
COST_INFRA       = [3600, 3960, 4440, 4920, 5400]
COST_PAYROLL     = [36000, 66000, 108000, 150000, 198000]
COST_PASSTHROUGH = [241, 314, 370, 445, 530]
COST_OTHER       = [9000, 14000, 20000, 26000, 32000]       # BD / travel / mktg
COST_TOTAL       = [48841, 84274, 132810, 181365, 235930]   # 5yr 683,220

# Bottom line (US$, exact from canon).
SURPLUS          = [12795, 46593, 65536, 95894, 125765]     # 5yr 346,583
CUM_SURPLUS      = [12795, 59388, 124924, 220818, 346583]
GMV_ON_RAILS     = [894600, 2463200, 3683300, 5310800, 7149100]  # 5yr 19,501,000

# Revenue mix (auction-house% / consumer-transport%).
MIX_HOUSE     = [0.987, 0.969, 0.952, 0.933, 0.912]
MIX_TRANSPORT = [0.013, 0.031, 0.048, 0.067, 0.088]

# ----------------------------------------------------------------------------
# Style helpers
# ----------------------------------------------------------------------------
TERRACOTTA = "B85042"; GOLD = "D4A843"; CREAM = "F5E6D3"
DARK = "2D1B1A"; MUTED = "7A4F47"; LIGHT_GREY = "F0F0EC"

thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(c):
    c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=TERRACOTTA)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border

def style_section(c):
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")

def style_input(c):
    c.font = Font(name="Calibri", size=10, color="000000")
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border

def style_calc(c):
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border

def style_total(c):
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=TERRACOTTA)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = border

def style_label(c):
    c.font = Font(name="Calibri", size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border

def style_label_bold(c):
    c.font = Font(name="Calibri", size=10, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = border

def fmt_currency(c): c.number_format = '"US$"#,##0;[Red]("US$"#,##0)'
def fmt_pct(c): c.number_format = "0.0%"
def fmt_pct2(c): c.number_format = "0.00%"
def fmt_int(c): c.number_format = "#,##0"

wb = openpyxl.Workbook()
wb.remove(wb.active)

YEAR_COLS = [get_column_letter(2 + i) for i in range(YEARS)]   # B..F
TOTAL_COL = get_column_letter(2 + YEARS)                       # G

# ============================================================================
# TAB 1 — ASSUMPTIONS
# ============================================================================
ws = wb.create_sheet("1. Assumptions")
for col, w in zip("ABCDE", (46, 16, 16, 16, 44)):
    ws.column_dimensions[col].width = w

ws.merge_cells("A1:E1")
ws["A1"] = "ZIMLIVESTOCK — FINANCIAL MODEL ASSUMPTIONS (v3.0, scaling B2B-SaaS platform)"
ws["A1"].font = Font(name="Georgia", size=13, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor=DARK)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32
ws.merge_cells("A2:E2")
ws["A2"] = ("Gold cells = inputs (edit these). White cells = calculations. "
            "Scaling case: ~20 houses over 5 years, field-honest adoption, small team, self-funded with a working-capital buffer.")
ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=MUTED)
ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
ws["A2"].fill = PatternFill("solid", fgColor=CREAM)

def section(row, label):
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = label
    style_section(ws[f"A{row}"])
    ws.row_dimensions[row].height = 20

def row_input(row, label, value, fmt="currency", note=""):
    ws[f"A{row}"] = label; style_label(ws[f"A{row}"])
    ws[f"B{row}"] = value; style_input(ws[f"B{row}"])
    if fmt == "currency": fmt_currency(ws[f"B{row}"])
    elif fmt == "pct": fmt_pct(ws[f"B{row}"])
    elif fmt == "pct2": fmt_pct2(ws[f"B{row}"])
    elif fmt == "int": fmt_int(ws[f"B{row}"])
    ws.merge_cells(f"C{row}:E{row}")
    ws[f"C{row}"] = note
    ws[f"C{row}"].font = Font(name="Calibri", size=9, italic=True, color=MUTED)
    ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center")

row = 4
section(row, "PRICING — onboarding fee (one-off, per house)"); row += 1
row_input(row, "Onboarding fee — Tier A (anchor house)", 3500, note="Self-serve wizard onboarding; ~6-min RLS-isolated tenant"); row += 1
row_input(row, "Onboarding fee — Tier B (mid-market)",   2500, note=""); row += 1
row_input(row, "Onboarding fee — Tier C (smaller house)", 1500, note=""); row += 1
row_input(row, "Onboarding fee — Pilot",                 1000, note="Credited toward the full tier fee on conversion (never paid twice)"); row += 1

row += 1
section(row, "PRICING — monthly platform subscription (per house)"); row += 1
row_input(row, "Subscription — Tier A",  1500, note="Recurring platform subscription"); row += 1
row_input(row, "Subscription — Tier B",  1200, note=""); row += 1
row_input(row, "Subscription — Tier C",   900, note=""); row += 1
row_input(row, "Subscription — Pilot",   1000, note="US$1,000/mo for 90 days, then converts to signed tier"); row += 1

row += 1
section(row, "PRICING — transaction take + transport"); row += 1
row_input(row, "Transaction take %",     0.0075, fmt="pct2", note="0.75% of settled GMV, on TOP of Paynow's fee"); row += 1
row_input(row, "Transport — base booking leg", 15, note="Buyer-facing flat fee; platform keeps only this leg"); row += 1
row_input(row, "Transport — per km (transporter keeps)", 0.35, note="Distance haul priced + kept by the transporter"); row += 1
row_input(row, "Transport — quote cap", 250, note="Caps the buyer quote"); row += 1
row_input(row, "Transport — platform NET per booking", 12.00, note="US$15 booking leg less ~US$3 processing"); row += 1

row += 1
section(row, "ADOPTION & DRIVERS — held field-honest (NOT a growth lever)"); row += 1
row_input(row, "Blended digital adoption — Year 1", 0.105, fmt="pct", note="Below the ~15% mature ceiling for a manual cash market"); row += 1
row_input(row, "Blended digital adoption — Year 5", 0.137, fmt="pct", note="Still field-honest at Y5"); row += 1
row_input(row, "Avg digital ticket", 650, note="Per settled digital transaction"); row += 1
row_input(row, "Transport attach-rate — Year 1", 0.05, fmt="pct", note="Ramps 5% -> 24% over Y1..Y5"); row += 1
row_input(row, "Transport attach-rate — Year 5", 0.24, fmt="pct", note="Consumer line diversifies revenue over time"); row += 1

row += 1
section(row, "SCALE & RAMP — houses live at year-end"); row += 1
row_input(row, "Houses live — Year 1", 5, fmt="int", note="Anchors-first; ~1/3 of the ~40-60 house market by Y5"); row += 1
row_input(row, "Houses live — Year 2", 8, fmt="int", note=""); row += 1
row_input(row, "Houses live — Year 3", 12, fmt="int", note=""); row += 1
row_input(row, "Houses live — Year 4", 16, fmt="int", note=""); row += 1
row_input(row, "Houses live — Year 5", 20, fmt="int", note="Subscription uses a half-year go-live convention"); row += 1

row += 1
section(row, "COSTS — small team, self-funded with a working-capital buffer"); row += 1
row_input(row, "Payroll — Year 1", 36000, note="Founder salary sits inside payroll"); row += 1
row_input(row, "Payroll — Year 5", 198000, note="Funds a real ~9-13 person org (not a lean 4-person team)"); row += 1
row_input(row, "Infrastructure — Year 1", 3600, note="Supabase + Vercel + Cloudflare + channels"); row += 1
row_input(row, "Other (BD / travel / mktg) — Year 1", 9000, note="Business development, travel, marketing"); row += 1

# ============================================================================
# TAB 2 — HOUSE RAMP
# ============================================================================
ws2 = wb.create_sheet("2. House ramp")
for col, w in zip("ABCDEFG", (30, 12, 12, 12, 12, 12, 14)):
    ws2.column_dimensions[col].width = w
ws2.merge_cells("A1:G1")
ws2["A1"] = "HOUSE RAMP — ~20 houses over five years (5 → 8 → 12 → 16 → 20 live, anchors-first)"
ws2["A1"].font = Font(name="Georgia", size=13, bold=True, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", fgColor=DARK)
ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 32

hdrs = ["", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
for i, h in enumerate(hdrs, start=1):
    style_header(ws2.cell(row=3, column=i, value=h))
ws2.row_dimensions[3].height = 24

def ramp_row(r, label, values, fmt="int", bold=False, total=False):
    cell = ws2.cell(row=r, column=1, value=label)
    (style_label_bold if bold else style_label)(cell)
    for i, v in enumerate(values):
        c = ws2.cell(row=r, column=2 + i, value=v)
        (style_total if total else style_calc)(c)
        if fmt == "int": fmt_int(c)
        elif fmt == "currency": fmt_currency(c)
        elif fmt == "pct": fmt_pct(c)

ramp_row(4, "Houses live (EOY)", HOUSES_LIVE, bold=True, total=True)
ramp_row(5, "  Tier A onboarded (year)", [a for (a, b, c) in ONBOARDING_ADDS])
ramp_row(6, "  Tier B onboarded (year)", [b for (a, b, c) in ONBOARDING_ADDS])
ramp_row(7, "  Tier C onboarded (year)", [c for (a, b, c) in ONBOARDING_ADDS])
ramp_row(8, "Cumulative Tier A", [a for (a, b, c) in TIER_MIX_CUM], bold=True)
ramp_row(9, "Cumulative Tier B", [b for (a, b, c) in TIER_MIX_CUM], bold=True)
ramp_row(10, "Cumulative Tier C", [c for (a, b, c) in TIER_MIX_CUM], bold=True)
ramp_row(12, "Blended digital adoption", BLENDED_ADOPTION, fmt="pct", bold=True)
ramp_row(13, "Digital transactions", DIGITAL_TRANSACTIONS, bold=True)
ramp_row(14, "Transport attach-rate", ATTACH_RATE, fmt="pct", bold=True)
ramp_row(16, "GMV onto Paynow rails", GMV_ON_RAILS, fmt="currency", bold=True, total=True)

note = ("Cumulative tier mix runs 3A/1B/1C (Y1) → 7A/8B/5C (Y5). Subscription uses a half-year "
        "go-live convention: the current-year cohort is billed 6 months, prior cohorts 12. "
        "Avg digital ticket US$650. Adoption stays field-honest (below the ~15% mature ceiling) — "
        "it is NOT used as a growth lever; growth rides house count + transport.")
ws2.merge_cells("A18:G20")
ws2["A18"] = note
ws2["A18"].font = Font(name="Calibri", size=9, italic=True, color=MUTED)
ws2["A18"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# ============================================================================
# Shared annual-table helpers for tabs 3/4/5
# ============================================================================
def annual_header(ws, row, first_label="", extra_total=True):
    style_header(ws.cell(row=row, column=1, value=first_label))
    for i in range(YEARS):
        style_header(ws.cell(row=row, column=2 + i, value=f"Year {i + 1}"))
    if extra_total:
        style_header(ws.cell(row=row, column=2 + YEARS, value="5-yr total"))
    ws.row_dimensions[row].height = 24

def annual_line(ws, row, label, values, *, bold=False, total_style=False,
                grand=False, fmt="currency", emit_total=True):
    cell = ws.cell(row=row, column=1, value=label)
    if grand:
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="left", vertical="center")
    else:
        (style_label_bold if bold else style_label)(cell)
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=2 + i, value=v)
        if grand:
            c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=DARK)
            c.alignment = Alignment(horizontal="right", vertical="center")
        else:
            (style_total if total_style else style_calc)(c)
        if fmt == "currency": fmt_currency(c)
        elif fmt == "pct": fmt_pct(c)
        elif fmt == "int": fmt_int(c)
    if emit_total:
        tc = ws.cell(row=row, column=2 + YEARS,
                     value=f"=SUM({YEAR_COLS[0]}{row}:{YEAR_COLS[-1]}{row})")
        if grand:
            tc.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            tc.fill = PatternFill("solid", fgColor=DARK)
            tc.alignment = Alignment(horizontal="right", vertical="center")
        else:
            (style_total if (total_style or bold) else style_calc)(tc)
        if fmt == "currency": fmt_currency(tc)
        elif fmt == "int": fmt_int(tc)

# ============================================================================
# TAB 3 — REVENUE BY YEAR (four lines)
# ============================================================================
ws3 = wb.create_sheet("3. Revenue")
for col, w in zip("ABCDEFG", (28, 15, 15, 15, 15, 15, 16)):
    ws3.column_dimensions[col].width = w
ws3.merge_cells("A1:G1")
ws3["A1"] = "ANNUAL REVENUE — 5-year projection across four lines"
ws3["A1"].font = Font(name="Georgia", size=13, bold=True, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", fgColor=DARK)
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

annual_header(ws3, 3, "Revenue line")
annual_line(ws3, 4, "Onboarding (one-off)", REV_ONBOARDING)
annual_line(ws3, 5, "Platform subscription", REV_SUBSCRIPTION)
annual_line(ws3, 6, "Transaction take (0.75%)", REV_TAKE)
annual_line(ws3, 7, "Transport (consumer line)", REV_TRANSPORT)
# Grand total computed from the four lines above.
ws3.cell(row=8, column=1, value="REVENUE TOTAL")
ws3.cell(row=8, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ws3.cell(row=8, column=1).fill = PatternFill("solid", fgColor=DARK)
for i in range(YEARS):
    col = YEAR_COLS[i]
    c = ws3.cell(row=8, column=2 + i, value=f"=SUM({col}4:{col}7)")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="right", vertical="center")
    fmt_currency(c)
gt = ws3.cell(row=8, column=2 + YEARS, value=f"=SUM({YEAR_COLS[0]}8:{YEAR_COLS[-1]}8)")
gt.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
gt.fill = PatternFill("solid", fgColor=DARK)
gt.alignment = Alignment(horizontal="right", vertical="center")
fmt_currency(gt)

# Revenue mix.
ws3.cell(row=10, column=1, value="REVENUE MIX")
ws3.cell(row=10, column=1).font = Font(name="Calibri", size=11, bold=True, color=TERRACOTTA)
annual_line(ws3, 11, "Auction-house revenue %", MIX_HOUSE, fmt="pct", emit_total=False)
annual_line(ws3, 12, "Consumer-transport revenue %", MIX_TRANSPORT, fmt="pct", emit_total=False)
ws3.cell(row=13, column=1, value="5-yr revenue total: US$1,029,803  ·  onboarding US$52,000 · subscription US$766,800 · take US$146,258 · transport US$64,745")
ws3.cell(row=13, column=1).font = Font(name="Calibri", size=9, italic=True, color=MUTED)
ws3.merge_cells("A13:G13")
ws3.freeze_panes = "B4"

# ============================================================================
# TAB 4 — COSTS BY YEAR (four lines)
# ============================================================================
ws4 = wb.create_sheet("4. Costs")
for col, w in zip("ABCDEFG", (28, 15, 15, 15, 15, 15, 16)):
    ws4.column_dimensions[col].width = w
ws4.merge_cells("A1:G1")
ws4["A1"] = "ANNUAL COSTS — small team + infrastructure (founder salary inside payroll)"
ws4["A1"].font = Font(name="Georgia", size=13, bold=True, color="FFFFFF")
ws4["A1"].fill = PatternFill("solid", fgColor=DARK)
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

annual_header(ws4, 3, "Cost line")
annual_line(ws4, 4, "Infrastructure", COST_INFRA)
annual_line(ws4, 5, "Payroll (incl. founder)", COST_PAYROLL)
annual_line(ws4, 6, "Pass-through (Paynow)", COST_PASSTHROUGH)
annual_line(ws4, 7, "Other (BD / travel / mktg)", COST_OTHER)
ws4.cell(row=8, column=1, value="COSTS TOTAL")
ws4.cell(row=8, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ws4.cell(row=8, column=1).fill = PatternFill("solid", fgColor=DARK)
for i in range(YEARS):
    col = YEAR_COLS[i]
    c = ws4.cell(row=8, column=2 + i, value=f"=SUM({col}4:{col}7)")
    c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=DARK)
    c.alignment = Alignment(horizontal="right", vertical="center")
    fmt_currency(c)
gt4 = ws4.cell(row=8, column=2 + YEARS, value=f"=SUM({YEAR_COLS[0]}8:{YEAR_COLS[-1]}8)")
gt4.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
gt4.fill = PatternFill("solid", fgColor=DARK)
gt4.alignment = Alignment(horizontal="right", vertical="center")
fmt_currency(gt4)
ws4.cell(row=10, column=1, value="5-yr costs total: US$683,220  ·  Y5 payroll US$198,000 funds a real ~9-13 person org (NOT a lean 4-person team).")
ws4.cell(row=10, column=1).font = Font(name="Calibri", size=9, italic=True, color=MUTED)
ws4.merge_cells("A10:G10")
ws4.freeze_panes = "B4"

# ============================================================================
# TAB 5 — P&L SUMMARY
# ============================================================================
ws5 = wb.create_sheet("5. P&L summary")
for col, w in zip("ABCDEFG", (28, 15, 15, 15, 15, 15, 16)):
    ws5.column_dimensions[col].width = w
ws5.merge_cells("A1:G1")
ws5["A1"] = "P&L SUMMARY — year by year (v3.0 scaling platform)"
ws5["A1"].font = Font(name="Georgia", size=13, bold=True, color="FFFFFF")
ws5["A1"].fill = PatternFill("solid", fgColor=DARK)
ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30

annual_header(ws5, 3, "")
# Revenue pulled from tab 3 grand total (row 8).
ws5.cell(row=4, column=1, value="Revenue"); style_label_bold(ws5.cell(row=4, column=1))
for i in range(YEARS):
    col = YEAR_COLS[i]
    c = ws5.cell(row=4, column=2 + i, value=f"='3. Revenue'!{col}8"); style_calc(c); fmt_currency(c)
tc = ws5.cell(row=4, column=2 + YEARS, value=f"=SUM({YEAR_COLS[0]}4:{YEAR_COLS[-1]}4)"); style_total(tc); fmt_currency(tc)

# Costs pulled from tab 4 grand total (row 8), shown negative.
ws5.cell(row=5, column=1, value="Costs"); style_label_bold(ws5.cell(row=5, column=1))
for i in range(YEARS):
    col = YEAR_COLS[i]
    c = ws5.cell(row=5, column=2 + i, value=f"=-'4. Costs'!{col}8"); style_calc(c); fmt_currency(c)
tc = ws5.cell(row=5, column=2 + YEARS, value=f"=SUM({YEAR_COLS[0]}5:{YEAR_COLS[-1]}5)"); style_total(tc); fmt_currency(tc)

# Surplus.
ws5.cell(row=6, column=1, value="SURPLUS")
ws5.cell(row=6, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
ws5.cell(row=6, column=1).fill = PatternFill("solid", fgColor=DARK)
for i in range(YEARS):
    col = YEAR_COLS[i]
    c = ws5.cell(row=6, column=2 + i, value=f"={col}4+{col}5")
    c.font = Font(name="Calibri", size=10, bold=True)
    c.fill = PatternFill("solid", fgColor=GOLD)
    c.alignment = Alignment(horizontal="right", vertical="center")
    fmt_currency(c)
sc = ws5.cell(row=6, column=2 + YEARS, value=f"=SUM({YEAR_COLS[0]}6:{YEAR_COLS[-1]}6)")
sc.font = Font(name="Calibri", size=10, bold=True)
sc.fill = PatternFill("solid", fgColor=GOLD)
sc.alignment = Alignment(horizontal="right", vertical="center")
fmt_currency(sc)

# Cumulative surplus (running).
ws5.cell(row=7, column=1, value="Cumulative surplus"); style_label_bold(ws5.cell(row=7, column=1))
for i in range(YEARS):
    col = YEAR_COLS[i]
    formula = f"={col}6" if i == 0 else f"={col}6+{YEAR_COLS[i-1]}7"
    c = ws5.cell(row=7, column=2 + i, value=formula); style_calc(c); fmt_currency(c)

# GMV onto rails.
annual_line(ws5, 8, "GMV onto Paynow rails", GMV_ON_RAILS, bold=True, total_style=True)
ws5.freeze_panes = "B4"

# Drivers block.
ws5.cell(row=10, column=1, value="DRIVERS")
ws5.cell(row=10, column=1).font = Font(name="Georgia", size=12, bold=True, color="FFFFFF")
ws5.cell(row=10, column=1).fill = PatternFill("solid", fgColor=DARK)
ws5.merge_cells("A10:G10")
for i in range(YEARS):
    style_header(ws5.cell(row=11, column=2 + i, value=f"Year {i + 1}"))
annual_line(ws5, 12, "Houses live (EOY)", HOUSES_LIVE, fmt="int", emit_total=False)
annual_line(ws5, 13, "Blended adoption", BLENDED_ADOPTION, fmt="pct", emit_total=False)
annual_line(ws5, 14, "Digital transactions", DIGITAL_TRANSACTIONS, fmt="int", emit_total=False)
annual_line(ws5, 15, "Transport attach-rate", ATTACH_RATE, fmt="pct", emit_total=False)
annual_line(ws5, 16, "Auction-house revenue %", MIX_HOUSE, fmt="pct", emit_total=False)
annual_line(ws5, 17, "Consumer-transport revenue %", MIX_TRANSPORT, fmt="pct", emit_total=False)

# Notes.
ws5.cell(row=19, column=1, value="NOTES — read before quoting any number")
ws5.cell(row=19, column=1).font = Font(name="Calibri", size=11, bold=True, color=TERRACOTTA)
ws5.merge_cells("A19:G19")
notes = [
    "SCALING B2B-SaaS PLATFORM. ~20 houses over 5 years (5 → 8 → 12 → 16 → 20 live, anchors-first), about 1/3 of the ~40-60 house market. A self-serve onboarding wizard (/operators → admin approval → ~6-min RLS-isolated tenant, no SQL) makes scale low-touch.",
    "Five live channels: web/PWA, WhatsApp, USSD, BillPay-as-biller, Facebook Messenger. B2B today (auction houses), B2B2C tomorrow (consumer transport diversifies revenue: 1.3% of revenue in Y1 → 8.8% in Y5).",
    "Positive surplus every year: +US$12,795 (Y1) → +US$125,765 (Y5); +US$346,583 cumulative over 5 years. No equity-requiring trough. The Y1 cushion is thin (~3.5 weeks of opex), so hold a 2-3 month opex working-capital buffer.",
    "Small team, largely SELF-FUNDED (NOT external equity). Y5 payroll US$198,000 funds a real ~9-13 person org — NOT a 'lean 4-person team'. The founder draws a salary that sits inside payroll.",
    "Subscription-led: Y5 subscription US$266,400 of US$361,695 total. Recurring spine is reliable; the 0.75% transaction take compounds with GMV; transport is the consumer diversification line.",
    "GMV routed onto Paynow rails: US$894,600 (Y1) → US$7,149,100 (Y5); ~US$19.5M over 5 years. Paynow is the settlement spine (Integration ID 23997; Stripe diaspora fallback). 3 of 4 national livestock-digitization initiatives have NO settlement layer — we are it.",
    "Field-honest caveats: digital adoption stays ~10.5-13.7% (below the ~15% mature ceiling) and is NOT used as a growth lever — growth rides house count + transport. The single MOST AGGRESSIVE assumption is landing 3 of ~8 Tier A anchors in Year 1. All figures USD.",
]
for i, n in enumerate(notes):
    ws5.cell(row=20 + i, column=1, value=f"•  {n}")
    ws5.cell(row=20 + i, column=1).font = Font(name="Calibri", size=10, color=MUTED)
    ws5.cell(row=20 + i, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws5.merge_cells(start_row=20 + i, start_column=1, end_row=20 + i, end_column=7)
    ws5.row_dimensions[20 + i].height = 32

# ============================================================================
# TAB 0 — README (cover)
# ============================================================================
cover = wb.create_sheet("0. README", 0)
cover.column_dimensions["A"].width = 100
cover["A1"] = "ZIMLIVESTOCK"
cover["A1"].font = Font(name="Georgia", size=24, bold=True, color="FFFFFF")
cover["A1"].fill = PatternFill("solid", fgColor=DARK)
cover["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
cover.row_dimensions[1].height = 50
cover["A2"] = "Financial Model — 5-year scaling B2B-SaaS platform (v3.0)"
cover["A2"].font = Font(name="Georgia", size=14, italic=True, color=GOLD)
cover["A2"].fill = PatternFill("solid", fgColor=DARK)
cover["A2"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
cover.row_dimensions[2].height = 28
cover["A3"] = "v3.0  ·  Tatenda Nyemudzo  ·  June 2026  ·  scaling platform (self-funded, working-capital buffer)"
cover["A3"].font = Font(name="Calibri", size=11, italic=True, color=MUTED)
cover["A3"].alignment = Alignment(horizontal="left", indent=1)
cover.row_dimensions[3].height = 24

readme_rows = [
    "",
    "WHAT'S IN THIS WORKBOOK",
    "",
    "1. Assumptions   — All inputs in gold cells. Pricing (4 lines), adoption, scale ramp, costs.",
    "2. House ramp    — Houses live by year (5 → 8 → 12 → 16 → 20), tier mix, GMV onto rails.",
    "3. Revenue       — Year-by-year across the 4 revenue lines, totaled, + revenue mix.",
    "4. Costs         — Year-by-year across the 4 cost lines (founder salary inside payroll).",
    "5. P&L summary   — Year-by-year revenue, costs, surplus, cumulative, GMV, drivers + notes.",
    "",
    "",
    "THE FRAME — A SCALING B2B-SaaS PLATFORM",
    "",
    "·  Auction houses ONBOARD as isolated (RLS) tenants via a self-serve wizard and pay a one-off",
    "   onboarding fee + a monthly platform subscription + a thin 0.75% take on settled GMV.",
    "·  B2B today; B2B2C tomorrow — the platform diversifies into consumer TRANSPORT (delivery).",
    "·  ~20 houses over 5 years (about 1/3 of the ~40-60 house market), anchors-first.",
    "·  A small team (~9-13 people by Y5), largely SELF-FUNDED with a 2-3 month working-capital",
    "   buffer (NOT external equity). The founder draws a salary inside payroll.",
    "·  Five live channels: web/PWA, WhatsApp, USSD, BillPay-as-biller, Facebook Messenger.",
    "",
    "",
    "HONEST HEADLINE OUTPUTS (v3.0)",
    "",
    "·  Year 1:  revenue US$61,636    cost US$48,841    surplus +US$12,795    (5 houses live)",
    "·  Year 3:  revenue US$198,346   cost US$132,810   surplus +US$65,536    (12 houses live)",
    "·  Year 5:  revenue US$361,695   cost US$235,930   surplus +US$125,765   (20 houses live)",
    "·  5-year:  revenue US$1,029,803   ·   surplus +US$346,583 cumulative.",
    "·  Positive surplus every year; no equity-requiring trough. Y1 cushion is thin (~3.5 weeks",
    "   of opex) — hold a 2-3 month opex working-capital buffer.",
    "·  GMV onto Paynow rails: US$894,600 → US$7,149,100 (~US$19.5M over 5 years).",
    "",
    "WHAT DRIVES GROWTH",
    "",
    "·  Growth rides HOUSE COUNT + TRANSPORT, NOT adoption. Adoption is held field-honest",
    "   (~10.5-13.7%, below the ~15% mature ceiling) and is not used as a growth lever.",
    "·  Subscription-led recurring spine; 0.75% take compounds with GMV; transport diversifies",
    "   revenue from 1.3% (Y1) to 8.8% (Y5) of the top line.",
    "",
    "HONEST CAVEATS",
    "",
    "·  The single MOST AGGRESSIVE assumption is landing 3 of ~8 Tier A anchors in Year 1.",
    "·  Y5 payroll US$198k funds a real ~9-13 person org — not a lean 4-person team.",
    "·  Paynow is the settlement spine (Integration ID 23997; Stripe diaspora fallback). 3 of 4",
    "   national livestock-digitization initiatives have no settlement layer — we are it.",
    "·  Zim-specific: USD scarcity / collection friction, currency volatility. All figures USD.",
]
section_titles = ("WHAT'S IN THIS WORKBOOK", "THE FRAME — A SCALING B2B-SaaS PLATFORM",
                  "HONEST HEADLINE OUTPUTS (v3.0)", "WHAT DRIVES GROWTH", "HONEST CAVEATS")
for i, txt in enumerate(readme_rows, start=5):
    cover[f"A{i}"] = txt
    if txt.strip() in section_titles:
        cover[f"A{i}"].font = Font(name="Calibri", size=11, bold=True, color=TERRACOTTA)
    else:
        cover[f"A{i}"].font = Font(name="Calibri", size=10, color=DARK)
    cover[f"A{i}"].alignment = Alignment(horizontal="left", indent=1, vertical="top")
    cover.row_dimensions[i].height = 16

wb.save("/Users/tatendanyemudzo/Downloads/app/deliverables/business/financial-model.xlsx")
print("Wrote: financial-model.xlsx")
